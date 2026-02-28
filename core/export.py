# core/export.py
from __future__ import annotations

import datetime as _dt
import os
from typing import List, Sequence, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def default_report_filename(title: str) -> str:
    ts = _dt.datetime.now().strftime("%Y-%m-%d_%H-%M")
    safe = "".join(ch if ch.isalnum() or ch in " _-." else "_" for ch in title).strip() or "report"
    return f"{safe}_{ts}.xlsx"


def export_table_to_xlsx(
    headers: Sequence[str],
    rows: Sequence[Sequence[str]],
    path: str,
    title: str = "Отчет",
    add_totals: bool = True
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"

    # Заголовок отчёта
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws["A2"] = f"Сформировано: {_dt.datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    start_row = 4

    # Шапка таблицы
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=str(h))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Данные
    r0 = start_row + 1
    for r, row in enumerate(rows, start=r0):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value="" if v is None else str(v))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    last_data_row = r0 + len(rows) - 1 if rows else start_row

    # Итого
    if add_totals:
        totals_row = last_data_row + 2
        ws.cell(row=totals_row, column=1, value=f"Итого: {len(rows)}").font = Font(bold=True)
        # красиво: объединяем на всю ширину таблицы
        if len(headers) >= 2:
            ws.merge_cells(
                start_row=totals_row, start_column=1,
                end_row=totals_row, end_column=len(headers)
            )

    # Автоширина колонок (по содержимому, но без фанатизма)
    max_width = 60
    for c in range(1, len(headers) + 1):
        col_letter = get_column_letter(c)
        best = 10
        best = max(best, len(str(headers[c - 1])) + 2)
        for r in range(r0, r0 + len(rows)):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            best = max(best, min(max_width, len(str(v)) + 2))
        ws.column_dimensions[col_letter].width = best

    # Заморозка шапки
    ws.freeze_panes = ws["A5"]  # строка после заголовка таблицы

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)
