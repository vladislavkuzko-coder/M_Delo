# core/import_lists.py
from __future__ import annotations

from typing import Tuple, List

from openpyxl import load_workbook

from core.db import upsert_dictionary_value


def import_txt_list(db_path: str, dict_type: str, filepath: str) -> Tuple[int, List[str]]:
    errors: List[str] = []
    n = 0
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            for line in f:
                v = (line or "").strip()
                if not v:
                    continue
                upsert_dictionary_value(db_path, dict_type, v, 1)
                n += 1
    except Exception as e:
        errors.append(str(e))
    return n, errors


def import_xlsx_list(db_path: str, dict_type: str, filepath: str) -> Tuple[int, List[str]]:
    errors: List[str] = []
    n = 0
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
            v = row[0]
            if v is None:
                continue
            v = str(v).strip()
            if not v:
                continue
            upsert_dictionary_value(db_path, dict_type, v, 1)
            n += 1
    except Exception as e:
        errors.append(str(e))
    return n, errors
