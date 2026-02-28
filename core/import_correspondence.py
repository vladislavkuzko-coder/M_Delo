from __future__ import annotations

from typing import Dict, Any, List, Tuple, Optional

import re
from openpyxl import load_workbook

import datetime as _dt

from core.db import connect
from core.correspondence import _dept_prefix
from core.numbering import extract_seq


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _detect_kind(sheet_name: str) -> Optional[str]:
    n = _norm(sheet_name)
    if "вход" in n or n.startswith("вх"):
        return "in"
    if "исход" in n or n.startswith("исх"):
        return "out"
    if "внутр" in n or n.startswith("вн"):
        return "internal"
    return None


def _find_header_row(rows: List[List[Any]]) -> Tuple[int, Dict[str, int]]:
    # scan first 30 rows to find a header containing known fields
    keys = {
        # ВАЖНО: не используем слишком общие варианты типа "номер"/"дата" для reg_no/reg_date,
        # иначе колонки «Номер/Дата» в исходящих попадают не туда.
        "reg_no": ["№ п/п", "№п/п", "№", "рег№", "рег №", "регистрационный номер", "id"],
        "reg_date": ["дата рег", "дата рег.", "дата рег. внутри", "дата рег. внутри отдела", "дата регистрации"],
        "doc_no": ["номер документа", "номер", "№ документа", "исх №", "doc_no", "номер вн.", "номер вн", "номер внутри"],
        "doc_date": ["дата документа", "дата", "doc_date", "дата вн.", "дата вн", "дата внутри"],
        "in_no": ["№ вх.", "№вх", "номер вх", "номер входящий", "in_no"],
        "in_date": ["дата вх.", "датавх", "дата входящая", "in_date"],
        "out_no": ["№ исх.", "№исх", "№ исх", "номер исх", "исх №", "out_no"],
        "out_date": ["дата исх.", "дата исх", "исх дата", "дата исходящая", "out_date"],
        "sender": ["отправитель", "sender"],
        "recipient": ["получатель", "адресат", "recipient"],
        "subject": ["содержание", "тема", "subject"],
        "due_date": ["срок", "срок ответа", "due", "due_date"],
        "executor": ["исполнитель", "executor"],
        "notes": ["примечание", "примечания", "notes"],
        "dept_reg_no": ["рег. № внутри", "рег № внутри", "рег. № внутри отдела", "рег№ внутри отдела", "dept_reg_no"],
        "dept_reg_date": ["дата рег. внутри", "дата рег внутри", "дата рег. внутри отдела", "дата рег внутри отдела", "dept_reg_date"],
        "work_state": ["статус", "состояние", "work_state"],
    }
    for ridx, row in enumerate(rows[:30]):
        m: Dict[str, int] = {}
        for cidx, v in enumerate(row):
            t = _norm(str(v)) if v is not None else ""
            if not t:
                continue
            for k, variants in keys.items():
                for vv in variants:
                    if vv in t:
                        m[k] = cidx
                        break
        # accept if we have enough signal that this is a header.
        # For outgoing/internal реестры users often have only "Дата" + "Номер" + parties + subject.
        # Header is accepted if we have enough signal.
        # IMPORTANT: outgoing registries may use headers "дата вх."/"номер вх." for columns that
        # should become "Дата"/"Номер" in the app. In that case, we detect them as in_date/in_no.
        if (
            ("reg_no" in m and len(m) >= 2)
            or ("reg_date" in m and len(m) >= 3)
            or ("doc_no" in m and "doc_date" in m)
            or ("in_no" in m and "in_date" in m)
            or ("out_no" in m and "out_date" in m)
            or ("dept_reg_no" in m and ("doc_date" in m or "reg_date" in m or "dept_reg_date" in m))
        ):
            # --- Fallback mapping for very simple exports ---
            # Частый кейс: в исходящих/внутренних листах заголовки только "Дата" и "Номер".
            # Мы намеренно не мапим слишком общие варианты в keys (см. комментарий выше),
            # но если строка уже признана заголовком — можем безопасно добавить эти поля.
            try:
                for cidx, v in enumerate(row):
                    t = _norm(str(v)) if v is not None else ""
                    if not t:
                        continue
                    if t == "дата":
                        if all(k not in m for k in ("doc_date", "out_date", "in_date", "dept_reg_date")):
                            m["doc_date"] = cidx
                    if t in ("номер", "номер документа") or t == "№":
                        # не перетираем явные рег./вх/исх/внутр номера
                        if all(k not in m for k in ("doc_no", "in_no", "out_no", "dept_reg_no")) and "reg_no" not in m:
                            m["doc_no"] = cidx
            except Exception:
                pass
            return ridx, m
    return 0, {}


def import_correspondence_xlsx(db_path: str, filepath: str) -> Tuple[int, int, List[str]]:
    """Import correspondence registry from XLSX.

    Возвращает (inserted, updated, errors). Сейчас обновление не делаем — только добавление,
    чтобы не терять данные из файла.
    """
    inserted = 0
    updated = 0
    errors: List[str] = []

    # NOTE: this import is used for relatively small registries, but we still
    # optimize DB writes because per-row connections/commits make it extremely
    # slow on some Windows/network drives.
    wb = load_workbook(filepath, data_only=True, read_only=True)

    con = connect(db_path, read_only=False)
    cur = con.cursor()
    try:
        con.execute("PRAGMA busy_timeout=20000;")
    except Exception:
        pass

    # Keep everything in a single transaction.
    try:
        cur.execute("BEGIN;")
    except Exception:
        pass

    # Caches
    max_reg_no: Dict[str, int] = {}
    dict_sort_next: Dict[str, int] = {}
    dict_seen: Dict[str, set[str]] = {
        "sender": set(),
        "recipient": set(),
        "executor": set(),
        # единый список организаций для договоров
        "lender": set(),
        "borrower": set(),
    }
    dept_seq_cache: Dict[Tuple[str, int], int] = {}

    def _max_reg(kind: str) -> int:
        if kind in max_reg_no:
            return max_reg_no[kind]
        try:
            cur.execute("SELECT COALESCE(MAX(reg_no),0) AS m FROM correspondence WHERE kind=?;", (kind,))
            r = cur.fetchone()
            if r is None:
                m = 0
            else:
                try:
                    m = int(r[0])
                except Exception:
                    m = int(r["m"])
        except Exception:
            try:
                cur.execute("SELECT COALESCE(MAX(reg_no),0) AS m FROM correspondence WHERE kind=?;", (kind,))
                m = int(cur.fetchone()["m"])
            except Exception:
                m = 0
        max_reg_no[kind] = int(m)
        return int(m)

    def _dict_upsert(dict_type: str, value: str) -> None:
        v = (value or "").strip()
        if not v:
            return
        # avoid repeating the same statement thousands of times
        s = dict_seen.get(dict_type)
        if s is not None and v in s:
            return
        try:
            if dict_type not in dict_sort_next:
                cur.execute("SELECT COALESCE(MAX(sort_order),0)+1 AS n FROM dictionary WHERE dict_type=?;", (dict_type,))
                dict_sort_next[dict_type] = int(cur.fetchone()["n"])
            so = int(dict_sort_next[dict_type])
            # Only 'active' is updated on conflict; sort_order is kept.
            cur.execute(
                """
                INSERT INTO dictionary(dict_type,value,active,sort_order)
                VALUES(?,?,?,?)
                ON CONFLICT(dict_type,value) DO UPDATE SET active=excluded.active;
                """,
                (dict_type, v, 1, so),
            )
            # We may have inserted a new row; bump sort_order for the next new value.
            # If it was a conflict-update, bumping is harmless (just creates gaps).
            dict_sort_next[dict_type] = so + 1
            if s is not None:
                s.add(v)
        except Exception:
            pass

    def _party_upsert(value: str) -> None:
        """Организация может быть и отправителем/получателем, и ссудодателем/ссудополучателем."""
        v = (value or "").strip()
        if not v:
            return
        _dict_upsert("lender", v)
        _dict_upsert("borrower", v)

    def _dept_next(kind: str, dept_reg_date: str) -> str:
        # legacy prefix (as used in create_item when cfg is absent)
        prefix = _dept_prefix(kind)
        if not prefix:
            return ""
        y = 0
        try:
            y = int((dept_reg_date or "")[:4])
        except Exception:
            y = _dt.date.today().year
        key = (kind, y)
        if key not in dept_seq_cache:
            # Find current max sequence for this year once.
            try:
                y0 = f"{y:04d}-01-01"
                y1 = f"{y:04d}-12-31"
                cur.execute(
                    """
                    SELECT dept_reg_no AS v FROM correspondence
                    WHERE kind=? AND COALESCE(dept_reg_date,'')!=''
                      AND substr(dept_reg_date,1,10) BETWEEN ? AND ?
                      AND dept_reg_no LIKE ?;
                    """,
                    (kind, y0, y1, f"{prefix}%"),
                )
                mx = 0
                for r in cur.fetchall():
                    v = (r[0] if not hasattr(r, "keys") else r["v"]) or ""
                    n = extract_seq(str(v), prefix, "")
                    if n and int(n) > mx:
                        mx = int(n)
                dept_seq_cache[key] = mx
            except Exception:
                dept_seq_cache[key] = 0
        dept_seq_cache[key] += 1
        seq = int(dept_seq_cache[key])
        return f"{prefix}{seq:04d}"

    ts = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for sh in wb.worksheets:
        kind = _detect_kind(sh.title) or None

        # Read first chunk for header detection (streaming for speed)
        it = sh.iter_rows(values_only=True)
        head_rows: List[List[Any]] = []
        try:
            for _i in range(40):
                head_rows.append(list(next(it)))
        except StopIteration:
            pass

        header_row, col = _find_header_row(head_rows)
        start = header_row + 1

        # if kind not detected, try by content: if has '№ вх' etc.
        if kind is None:
            if "in_no" in col or "in_date" in col:
                kind = "in"
            elif "dept_reg_no" in col:
                kind = "internal"
            else:
                kind = "out"

        # --- Header remap for user-friendly outgoing/internal exports ---
        # Пользователь может называть колонки "дата вх."/"номер вх." даже в ИСХОДЯЩИХ,
        # а в программе это должно лечь в "Дата"/"Номер" (doc_date/doc_no).
        if kind == "out":
            if "doc_date" not in col and "in_date" in col:
                col["doc_date"] = col["in_date"]
            if "doc_no" not in col and "in_no" in col:
                col["doc_no"] = col["in_no"]
        if kind == "internal":
            # иногда внутренние тоже отдают как "дата" без уточнений; doc_date у нас есть fallback.
            pass

        def _as_date_str(v: Any) -> str:
            """Нормализует дату к YYYY-MM-DD (без времени)."""
            if v is None:
                return ""
            import datetime as _dt
            if isinstance(v, _dt.datetime):
                return v.date().isoformat()
            if isinstance(v, _dt.date):
                return v.isoformat()
            s = str(v).strip()
            if not s:
                return ""
            # ISO datetime -> ISO date
            m = re.match(r"^(\d{4}-\d{2}-\d{2})\s+\d{2}:\d{2}:\d{2}", s)
            if m:
                return m.group(1)
            # DD.MM.YYYY -> ISO
            m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", s)
            if m:
                return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
            # already ISO date
            m = re.match(r"^(\d{4}-\d{2}-\d{2})$", s)
            if m:
                return s
            return s

        def get(row: List[Any], key: str) -> str:
            if key not in col:
                return ""
            v = row[col[key]]
            # даты приводим к нормальному виду без времени
            if key in ("reg_date", "doc_date", "out_date", "in_date", "due_date"):
                return _as_date_str(v)
            if key in ("dept_reg_date",):
                return _as_date_str(v)
            return "" if v is None else str(v).strip()

        def _norm_state(v: str) -> str:
            vv = _norm(v)
            if vv in ("исполнено", "исполнен", "закрыто", "закрыт", "done", "выполнено"):
                return "done"
            return "in_work"
        # Process buffered rows after header
        ridx = start
        buffered = head_rows[start:]
        for row in buffered:
            if not any((str(x).strip() if x is not None else "") for x in row):
                ridx += 1
                continue
            # иногда после таблицы есть "пустая" строка с одной цифрой/мусором — пропускаем
            try:
                tail = "".join((str(x).strip() for x in row[1:] if x is not None))
                if not tail.strip():
                    continue
            except Exception:
                pass

            # reg_no: prefer numeric first col
            reg_no = 0
            raw_reg = get(row, "reg_no") or (str(row[0]).strip() if len(row) else "")
            try:
                reg_no = int(float(raw_reg)) if raw_reg not in ("", None) else 0
            except Exception:
                reg_no = 0

            data: Dict[str, Any] = {"kind": kind}
            if reg_no > 0:
                data["reg_no"] = reg_no

            # common fields
            data["reg_date"] = get(row, "reg_date")
            data["sender"] = get(row, "sender")
            data["recipient"] = get(row, "recipient")
            data["subject"] = get(row, "subject")
            data["notes"] = get(row, "notes")
            data["due_date"] = get(row, "due_date")
            data["executor"] = get(row, "executor")


            # incoming work state (UI column "Статус")
            if kind == "in":
                raw_state = get(row, "work_state")
                if raw_state:
                    data["work_state"] = _norm_state(raw_state)

            # kind-specific
            if kind == "in":
                data["out_date"] = get(row, "out_date") or get(row, "doc_date")
                data["out_no"] = get(row, "out_no") or get(row, "doc_no")
                data["in_date"] = get(row, "in_date") or get(row, "doc_date")
                data["in_no"] = get(row, "in_no") or get(row, "doc_no")
                # dept_reg_date for incoming is "Дата рег. внутри" in UI
                data["dept_reg_date"] = get(row, "dept_reg_date") or get(row, "reg_date")
            elif kind == "out":
                # В некоторых Excel-файлах для исходящих используют заголовки "дата вх."/"номер вх.".
                # В приложении эти поля называются "Дата" и "Номер".
                data["doc_date"] = get(row, "doc_date") or get(row, "in_date") or get(row, "reg_date")
                data["doc_no"] = get(row, "doc_no") or get(row, "in_no")
            else:  # internal
                data["doc_date"] = get(row, "doc_date") or get(row, "reg_date")
                data["dept_reg_no"] = get(row, "dept_reg_no") or get(row, "doc_no")
                data["dept_reg_date"] = get(row, "dept_reg_date") or get(row, "reg_date")

            try:
                # reg_no fallback (keep ordering if not provided)
                if int(data.get("reg_no") or 0) <= 0:
                    mx = _max_reg(kind)
                    mx += 1
                    max_reg_no[kind] = mx
                    data["reg_no"] = mx

                # Dictionaries for autocomplete
                if data.get("sender"):
                    sv = str(data.get("sender") or "")
                    _dict_upsert("sender", sv)
                    _party_upsert(sv)
                if data.get("recipient"):
                    rv = str(data.get("recipient") or "")
                    _dict_upsert("recipient", rv)
                    _party_upsert(rv)
                if data.get("executor"):
                    _dict_upsert("executor", str(data.get("executor") or ""))

                # Work state + done_date
                work_state = (data.get("work_state") or "in_work") if kind == "in" else "in_work"
                done_date = (data.get("done_date") or "").strip()
                if kind == "in" and work_state == "done" and not done_date:
                    done_date = _dt.date.today().isoformat()
                if kind != "in":
                    done_date = ""

                # Dept numbering (legacy) for incoming/internal when empty
                if kind in ("in", "internal"):
                    dept_reg_date = (data.get("dept_reg_date") or "").strip() or _dt.date.today().isoformat()
                    data["dept_reg_date"] = dept_reg_date
                    if not (data.get("dept_reg_no") or "").strip():
                        data["dept_reg_no"] = _dept_next(kind, dept_reg_date)

                cur.execute(
                    """
                    INSERT INTO correspondence(
                        kind, reg_no, reg_date,
                        sender, recipient,
                        subject, notes, attachments,
                        ref_in_no, doc_date, doc_no, in_date, in_no, out_date, out_no,
                        status, due_date, executor, dept_reg_no, dept_reg_date,
                        work_state, done_date,
                        created_ts, updated_ts,
                        owner_unit_id, created_by
                    )
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);
                    """,
                    (
                        kind,
                        int(data.get("reg_no") or 0),
                        (data.get("reg_date") or "").strip(),
                        (data.get("sender") or "").strip(),
                        (data.get("recipient") or "").strip(),
                        (data.get("subject") or "").strip(),
                        (data.get("notes") or "").strip(),
                        (data.get("attachments") or "").strip(),
                        (data.get("ref_in_no") or "").strip(),
                        (data.get("doc_date") or "").strip(),
                        (data.get("doc_no") or "").strip(),
                        (data.get("in_date") or "").strip(),
                        (data.get("in_no") or "").strip(),
                        (data.get("out_date") or "").strip(),
                        (data.get("out_no") or "").strip(),
                        (data.get("status") or "").strip(),
                        (data.get("due_date") or "").strip(),
                        (data.get("executor") or "").strip(),
                        (data.get("dept_reg_no") or "").strip(),
                        (data.get("dept_reg_date") or "").strip(),
                        work_state,
                        done_date,
                        ts,
                        ts,
                        None,
                        "",
                    ),
                )
                inserted += 1
            except Exception as e:
                errors.append(f"{sh.title}: строка {ridx+1}: {e}")
            ridx += 1

        # Process remaining rows from streaming iterator
        for row in it:
            row = list(row)
            if not any((str(x).strip() if x is not None else "") for x in row):
                ridx += 1
                continue
            try:
                tail = "".join((str(x).strip() for x in row[1:] if x is not None))
                if not tail.strip():
                    ridx += 1
                    continue
            except Exception:
                pass

            # reuse existing parsing logic by duplicating the core section
            reg_no = 0
            raw_reg = get(row, "reg_no") or (str(row[0]).strip() if len(row) else "")
            try:
                reg_no = int(float(raw_reg)) if raw_reg not in ("", None) else 0
            except Exception:
                reg_no = 0

            data: Dict[str, Any] = {"kind": kind}
            if reg_no > 0:
                data["reg_no"] = reg_no

            data["reg_date"] = get(row, "reg_date")
            data["sender"] = get(row, "sender")
            data["recipient"] = get(row, "recipient")
            data["subject"] = get(row, "subject")
            data["notes"] = get(row, "notes")
            data["due_date"] = get(row, "due_date")
            data["executor"] = get(row, "executor")

            if kind == "in":
                raw_state = get(row, "work_state")
                if raw_state:
                    data["work_state"] = _norm_state(raw_state)

            if kind == "in":
                data["out_date"] = get(row, "out_date") or get(row, "doc_date")
                data["out_no"] = get(row, "out_no") or get(row, "doc_no")
                data["in_date"] = get(row, "in_date") or get(row, "doc_date")
                data["in_no"] = get(row, "in_no") or get(row, "doc_no")
                data["dept_reg_date"] = get(row, "dept_reg_date") or get(row, "reg_date")
            elif kind == "out":
                data["doc_date"] = get(row, "doc_date") or get(row, "in_date") or get(row, "reg_date")
                data["doc_no"] = get(row, "doc_no") or get(row, "in_no")
            else:
                data["doc_date"] = get(row, "doc_date") or get(row, "reg_date")
                data["dept_reg_no"] = get(row, "dept_reg_no") or get(row, "doc_no")
                data["dept_reg_date"] = get(row, "dept_reg_date") or get(row, "reg_date")

            try:
                if int(data.get("reg_no") or 0) <= 0:
                    mx = _max_reg(kind)
                    mx += 1
                    max_reg_no[kind] = mx
                    data["reg_no"] = mx

                if data.get("sender"):
                    sv = str(data.get("sender") or "")
                    _dict_upsert("sender", sv)
                    _party_upsert(sv)
                if data.get("recipient"):
                    rv = str(data.get("recipient") or "")
                    _dict_upsert("recipient", rv)
                    _party_upsert(rv)
                if data.get("executor"):
                    _dict_upsert("executor", str(data.get("executor") or ""))

                work_state = (data.get("work_state") or "in_work") if kind == "in" else "in_work"
                done_date = (data.get("done_date") or "").strip()
                if kind == "in" and work_state == "done" and not done_date:
                    done_date = _dt.date.today().isoformat()
                if kind != "in":
                    done_date = ""

                if kind in ("in", "internal"):
                    dept_reg_date = (data.get("dept_reg_date") or "").strip() or _dt.date.today().isoformat()
                    data["dept_reg_date"] = dept_reg_date
                    if not (data.get("dept_reg_no") or "").strip():
                        data["dept_reg_no"] = _dept_next(kind, dept_reg_date)

                cur.execute(
                    """
                    INSERT INTO correspondence(
                        kind, reg_no, reg_date,
                        sender, recipient,
                        subject, notes, attachments,
                        ref_in_no, doc_date, doc_no, in_date, in_no, out_date, out_no,
                        status, due_date, executor, dept_reg_no, dept_reg_date,
                        work_state, done_date,
                        created_ts, updated_ts,
                        owner_unit_id, created_by
                    )
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?);
                    """,
                    (
                        kind,
                        int(data.get("reg_no") or 0),
                        (data.get("reg_date") or "").strip(),
                        (data.get("sender") or "").strip(),
                        (data.get("recipient") or "").strip(),
                        (data.get("subject") or "").strip(),
                        (data.get("notes") or "").strip(),
                        (data.get("attachments") or "").strip(),
                        (data.get("ref_in_no") or "").strip(),
                        (data.get("doc_date") or "").strip(),
                        (data.get("doc_no") or "").strip(),
                        (data.get("in_date") or "").strip(),
                        (data.get("in_no") or "").strip(),
                        (data.get("out_date") or "").strip(),
                        (data.get("out_no") or "").strip(),
                        (data.get("status") or "").strip(),
                        (data.get("due_date") or "").strip(),
                        (data.get("executor") or "").strip(),
                        (data.get("dept_reg_no") or "").strip(),
                        (data.get("dept_reg_date") or "").strip(),
                        work_state,
                        done_date,
                        ts,
                        ts,
                        None,
                        "",
                    ),
                )
                inserted += 1
            except Exception as e:
                errors.append(f"{sh.title}: строка {ridx+1}: {e}")
            ridx += 1

    try:
        con.commit()
    except Exception:
        pass
    try:
        con.close()
    except Exception:
        pass

    return inserted, updated, errors